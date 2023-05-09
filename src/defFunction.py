from openpyxl import Workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.wait import WebDriverWait
from locators import HomePageLocators
import openpyxl
import pyperclip
from selenium.webdriver.chrome.options import Options

file_name = "F:\\upsc_questions.xlsx"  # Note the double backslashes or use forward slashes


class PoeSite:
    def __init__(self):
        # Set up Chrome driver
        self.clipboard = None
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)

    def navigate_to_website(self, url):
        # Navigate to the website
        self.driver.get(url)

    def click_use_email_button(self):
        self.driver.find_element(*HomePageLocators.email_but).click()

    def click_on_email_box(self):
        # locate email input box
        self.driver.find_element(*HomePageLocators.email_input).click()
        self.driver.find_element(*HomePageLocators.email_input).send_keys("shidorifoodsonline@gmail.com")

    def click_go_button(self):
        # Locate and click the go button
        self.driver.find_element(*HomePageLocators.go_butt).click()

    def click_login_button(self):
        self.driver.find_element(*HomePageLocators.login_butt).click()

    def wait(self, seconds):
        time.sleep(1)

    def close_browser(self):
        self.driver.quit()

    def ask_questions(self, chrome_options=True):
        # Load the Excel file
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook['Sheet1']
        print("Opening excel file" + file_name)
        max_row = worksheet.max_row
        self.driver.find_element(*HomePageLocators.question_area).click()

        # Loop through each row in the Excel file
        for row_number in range(1, max_row + 1):
            try:
                print(" entering in loop...")
                # past the question and click to send
                self.put_question(row_number, worksheet)
                self.driver.implicitly_wait(3)

                # Wait for the like button to be visible
                index = self.wait_like_button(row_number)

                time.sleep(2)
                # Wait for the hover to Ans box
                self.hover_ans_box(index)

                time.sleep(1)
                # click to drop down menu
                self.click_drop_down(index)


                time.sleep(1)
                # Click on the copy button
                self.click_copy_butt(index)
                print(" Copying the data..")

                time.sleep(1)
                # Copy data from clipboard
                clipboard, copied_data = self.copy_clipboard()

                # Write data to Excel file
                self.save_ans(copied_data, row_number, worksheet)


                time.sleep(1)
                # Save the Excel file
                workbook.save(file_name)
                print("saving the answer in column 3 and row " + str(row_number) + " to the file")
                # Reset clipboard
                clipboard.copy('')


                time.sleep(1)
                #  copy and click on suggest question 4th question
                self.copy_question(index, row_number, worksheet)
                self.click_suggest_question(index)
                print(" clicking the suggestion question")

                time.sleep(1)
                # Wait for the  suggest question ans like button to be visible
                index2 = self.wait_sugg_like(index)

                # Wait for the suggest Ans box to hover
                self.hover_sugg_ans_box(index2)

                # click to  suggest drop down menu
                self.click_sugg_drop_down(index2)

                # Click on the copy button
                self.click_sugg_copy_butt(index2)
                print("copying text....")


                time.sleep(1)
                clipboard, copied_data = self.copy_clipboard()

                time.sleep(1)
                # Write data to Excel file
                self.save_sugg_ans(copied_data, row_number, worksheet)

                time.sleep(1)
                # Save the Excel file
                workbook.save(file_name)
                print("saving the  suggested answer in column 4 and row " + str(row_number) + " to the file")


                # Reset clipboard
                clipboard.copy('')

            except Exception as e:
                print("Error occurred for row " + str(row_number) + ": " + str(e))
                print("------********  ERROR *********---------- ")

        print("Saving All the answers of Questions " + str(max_row) + " in the file name " + file_name)

    def save_ans(self, copied_data, row_number, worksheet):
        worksheet.cell(row=row_number, column=2).value = copied_data
        print("Saving the answer to row " + str(row_number))

    def copy_clipboard(self):
        clipboard = pyperclip
        copied_data = clipboard.paste()
        return clipboard, copied_data

    def click_copy_butt(self, index):
        wait = WebDriverWait(self.driver, timeout=20)
        copy_but = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[2]/div[3]/div/div/div/div/button[3]".format(
            index)
        copy = wait.until(EC.visibility_of_element_located((By.XPATH, copy_but)))
        time.sleep(2)
        copy.click()
        print("copying text....")

    def click_drop_down(self, index):
        time.sleep(1)
        wait = WebDriverWait(self.driver, timeout=20)
        hover_xpath = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[2]/div[3]/div/div/button/div".format(index)
        hover = wait.until(EC.visibility_of_element_located((By.XPATH, hover_xpath)))
        time.sleep(1)
        actions = ActionChains(self.driver)
        actions.move_to_element(hover).click().perform()

    def hover_ans_box(self, index):
        wait = WebDriverWait(self.driver, timeout=20)
        hover_box = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[2]/div[2]/div[1]/div[1]".format(index)
        box = wait.until(EC.visibility_of_element_located((By.XPATH, hover_box)))
        actions = ActionChains(self.driver)
        actions.move_to_element(box).click().perform()

    def wait_like_button(self, row_number):
        index = row_number + (row_number - 1)
        wait = WebDriverWait(self.driver, timeout=60)
        like_button = "//*[@id='__next']/div[1]/section/div/div/div[{}]/section/button[2]".format(index)
        wait.until(EC.visibility_of_element_located((By.XPATH, like_button)))
        return index

    def put_question(self, row_number, worksheet):
        question = worksheet.cell(row=row_number, column=1).value
        print("Asking question no " + str(row_number) + " to Sage")
        # Enter the question in the text area and click on send
        self.driver.find_element(*HomePageLocators.question_area).send_keys(question)
        self.driver.find_element(*HomePageLocators.que_send_butt).click()

    def copy_question(self, index, row_number, worksheet):
        button_xpath = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[3]/section/button[4]".format(index)
        wait = WebDriverWait(self.driver, timeout=10)
        wait.until(EC.visibility_of_element_located((By.XPATH, button_xpath)))
        copy_ques = self.driver.find_element(By.XPATH, button_xpath).text
        worksheet.cell(row=row_number, column=3).value = copy_ques
        print("questioned copied and save column 3 and row " + str(row_number) + " " + file_name)

    def click_suggest_question(self, index,):
        button_xpath = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[3]/section/button[4]".format(index)
        self.driver.find_element(By.XPATH, button_xpath).click()

    def hover_sugg_ans_box(self, index2):
        wait = WebDriverWait(self.driver, timeout=20)
        hover_sugg_ansbox = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[2]/div[2]/div[1]/div[1]".format(index2)
        box = wait.until(EC.visibility_of_element_located((By.XPATH, hover_sugg_ansbox)))
        actions = ActionChains(self.driver)
        actions.move_to_element(box).click().perform()

    def wait_sugg_like(self, index):
        index2 = index + 1
        wait = WebDriverWait(self.driver, timeout=60)
        like_button = "//*[@id='__next']/div[1]/section/div/div/div[{}]/section/button[2]".format(index2)
        wait.until(EC.visibility_of_element_located((By.XPATH, like_button)))
        return index2

    def click_sugg_drop_down(self, index2):
        time.sleep(2)
        wait = WebDriverWait(self.driver, timeout=20)
        hover_xpath = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[2]/div[3]/div/div/button/div".format(index2)
        hover = wait.until(EC.visibility_of_element_located((By.XPATH, hover_xpath)))
        time.sleep(2)
        actions = ActionChains(self.driver)
        actions.move_to_element(hover).click().perform()

    def click_sugg_copy_butt(self, index2):
        wait = WebDriverWait(self.driver, timeout=20)
        copy_but = "//*[@id='__next']/div[1]/section/div/div/div[{}]/div[2]/div[3]/div/div/div/div/button[3]".format(index2)
        copy = wait.until(EC.visibility_of_element_located((By.XPATH, copy_but)))
        time.sleep(1)
        copy.click()

    def save_sugg_ans(self, copied_data, row_number, worksheet):
        worksheet.cell(row=row_number, column=4).value = copied_data
        print("Saving the answer to row " + str(row_number))